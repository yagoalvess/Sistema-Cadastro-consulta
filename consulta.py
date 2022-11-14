import openpyxl
import sqlite3



def SalvarDadosExcel():
    #Criar a planilha
    book = openpyxl.Workbook()

    #Ver paginas existentes
    print(book.sheetnames)

    #Como criar uma pagina
    book.create_sheet('Usuarios')


    #Como selecionar uma pagina
    frutas_page = book['Usuarios']

    caminho = "D:\\Documents\\Desafio Tech\\db.sqlite3"
    con = sqlite3.connect(caminho)
    cursor= con.cursor()
    cursor.execute("select * from app_cliente")

    for linha in cursor:
        print(linha)
        frutas_page.append(linha)


    #salvar a planilha
    book.save('Consulta de usuarios.xlsx')

SalvarDadosExcel()